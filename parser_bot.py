import openpyxl
import json

# Загрузка рабочей книги Excel
workbook = openpyxl.load_workbook('06_05_2024-11_05_2024.xlsx')

# Выбор активного листа
sheet = workbook.active

# Словарь для хранения данных
data = {}

# Проход по группам
for col in range(5, 105, 2):  # E5 to DA5 with step 2 because room number is in the next column
    group = sheet.cell(row=5, column=col).value
    if group and group.endswith("9"):  # Если название группы заканчивается на "-9"
        group = group.replace("-9", "")  # Удалить "-9" из названия группы
    data[group] = {}

    # Проход по дням недели
    for row in range(7, 128):  # A7 to A127
        day = sheet.cell(row=row, column=1).value
        if day:  # Only add the day if it's not empty
            date = sheet.cell(row=row, column=2).value
            data[group][day] = {"date": date, "lessons": []}

            # Проход по парам
            for lesson_number in range(1, 8):  # There are 7 lessons in a day
                lesson_row = row + (lesson_number-1)*3  # Each lesson takes 3 rows
                lesson_time = sheet.cell(row=lesson_row, column=4).value
                lesson_name = sheet.cell(row=lesson_row, column=col).value
                teacher_name = sheet.cell(row=lesson_row+2, column=col).value
                room_number = sheet.cell(row=lesson_row, column=col+1).value  # Room number is in the next column

                if not lesson_name:  # Если ячейка с названием пары пуста
                    lesson_name = "Пары нет"
                    teacher_name = ""
                    room_number = ""

                lesson = {
                    "number": lesson_number,
                    "time": lesson_time,
                    "name": lesson_name,
                    "teacher": teacher_name,
                    "room": room_number,
                }
                data[group][day]["lessons"].append(lesson)

# Преобразование данных в JSON
json_data = json.dumps(data, ensure_ascii=False, indent=4)

# Запись JSON данных в файл
with open('rasp-data.json', 'w', encoding='utf-8') as f:
    f.write(json_data)
